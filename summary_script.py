
import pandas as pd
import numpy as np
from scipy.stats import pearsonr
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings
import os
import sys
import re

warnings.filterwarnings('ignore')


METRICS_TO_NORMALIZE = {
    'Pylint': [
        'Total Issues',
        'Convention (C)',
        'Refactor (R)',
        'Warning (W)',
        'Error (E)',
        'Fatal (F)'
    ],
    'Bandit': [
        'Codes With Issues',
        'Total Issues',
        'High Severity',
        'Medium Severity',
        'Low Severity',
        'Undefined Severity',
        'High Confidence',
        'Medium Confidence',
        'Low Confidence'
    ],
    'Radon': [
        'Total Complexity',
        'Total LOC',
        'Total LLOC',
        'Total SLOC',
        'Total Comments',
        'Total Blank',
        'Total Functions',
        'Grade A',
        'Grade B',
        'Grade C',
        'Grade D',
        'Grade E',
        'Grade F'
    ],
    'Sonarqube': [
        'Total Issues',
        'Severity: Blocker',
        'Severity: Critical',
        'Severity: Major',
        'Severity: Minor',
        'Severity: Info',
        'Type: Bug',
        'Type: Vulnerability',
        'Type: Code Smell',
        'Type: Security Hotspot',
        'Attribute: Consistency',
        'Attribute: Intentionality',
        'Attribute: Adaptability',
        'Attribute: Responsibility',
        'SQ Security: Blocker',
        'SQ Security: High',
        'SQ Security: Medium',
        'SQ Security: Low',
        'SQ Security: Info',
        'SQ Reliability: Blocker',
        'SQ Reliability: High',
        'SQ Reliability: Medium',
        'SQ Reliability: Low',
        'SQ Reliability: Info',
        'SQ Maintainability: Blocker',
        'SQ Maintainability: High',
        'SQ Maintainability: Medium',
        'SQ Maintainability: Low',
        'SQ Maintainability: Info'
    ]
}


METRICS_ALREADY_NORMALIZED = {
    'Pylint': ['Avg Issues Per Code', 'Avg Score', 'Correctness Rate (%)'],
    'Bandit': ['Issue Rate (%)', 'Avg Issues Per Code', 'Correctness Rate (%)'],
    'Radon': ['Avg Complexity', 'Avg Maintainability Index', 'Avg LOC', 'Correctness Rate (%)'],
    'Sonarqube': ['Avg Issues Per Code', 'Correctness Rate (%)',
                  'Avg Reliability Issues Per Code', 'Avg Maintainability Issues Per Code',
                  'Avg Security Issues Per Code']
}


def extract_difficulty_perfect(model_name, all_model_names):

    model_stripped = model_name.strip()


    difficulty_patterns = [
        (r'\(easy\)$', 'Easy'),
        (r'\(medium\)$', 'Medium'),
        (r'\(hard\)$', 'Hard')
    ]

    for pattern, difficulty in difficulty_patterns:
        if re.search(pattern, model_stripped, re.IGNORECASE):

            base_name = re.sub(pattern, '', model_stripped, flags=re.IGNORECASE).strip()


            if base_name in all_model_names:
                return difficulty


            has_difficulty_versions = False
            for d in ['Easy', 'Medium', 'Hard', 'easy', 'medium', 'hard']:
                potential_version = f"{model_stripped} ({d})"
                if potential_version in all_model_names:
                    has_difficulty_versions = True
                    break

            if has_difficulty_versions:

                return 'All'
            else:

                return 'All'


    return 'All'


def normalize_metrics(df, tool_name):

    df_normalized = df.copy()


    metrics_to_norm = METRICS_TO_NORMALIZE.get(tool_name, [])

    normalized_count = 0
    skipped_count = 0


    for metric in metrics_to_norm:
        if metric in df_normalized.columns:

            new_col_name = f"{metric} (per code)"



            total_codes = df_normalized['Total Codes'].values
            metric_values = df_normalized[metric].values

            normalized_values = np.where(
                total_codes > 0,
                metric_values / total_codes,
                0
            )


            df_normalized[new_col_name] = normalized_values

            normalized_count += 1
            print(f"      {metric}  {new_col_name}")
        else:
            skipped_count += 1
            print(f"      '{metric}' ")

    if normalized_count > 0:
        print(f"    {normalized_count} ")

    if skipped_count > 0:
        print(f"     {skipped_count} ")

    return df_normalized


def add_sonarqube_aggregated_metrics(df):

    df_enhanced = df.copy()


    reliability_cols = [
        'SQ Reliability: Blocker',
        'SQ Reliability: High',
        'SQ Reliability: Medium',
        'SQ Reliability: Low',
        'SQ Reliability: Info'
    ]

    maintainability_cols = [
        'SQ Maintainability: Blocker',
        'SQ Maintainability: High',
        'SQ Maintainability: Medium',
        'SQ Maintainability: Low',
        'SQ Maintainability: Info'
    ]

    security_cols = [
        'SQ Security: Blocker',
        'SQ Security: High',
        'SQ Security: Medium',
        'SQ Security: Low',
        'SQ Security: Info'
    ]


    total_codes = df_enhanced['Total Codes'].values


    reliability_sum = df_enhanced[reliability_cols].sum(axis=1).values
    df_enhanced['Avg Reliability Issues Per Code'] = np.where(
        total_codes > 0,
        reliability_sum / total_codes,
        0
    )


    maintainability_sum = df_enhanced[maintainability_cols].sum(axis=1).values
    df_enhanced['Avg Maintainability Issues Per Code'] = np.where(
        total_codes > 0,
        maintainability_sum / total_codes,
        0
    )


    security_sum = df_enhanced[security_cols].sum(axis=1).values
    df_enhanced['Avg Security Issues Per Code'] = np.where(
        total_codes > 0,
        security_sum / total_codes,
        0
    )

    print(f"   : Avg Reliability Issues Per Code")
    print(f"   : Avg Maintainability Issues Per Code")
    print(f"   : Avg Security Issues Per Code")

    return df_enhanced


def calculate_correlation(x, y):


    if len(x) < 3 or len(y) < 3:
        return None

    if np.std(x) == 0 or np.std(y) == 0:
        return None


    mask = ~(np.isnan(x) | np.isnan(y) | np.isinf(x) | np.isinf(y))
    x_clean = x[mask]
    y_clean = y[mask]

    if len(x_clean) < 3:
        return None

    try:

        pearson_corr, pearson_p = pearsonr(x_clean, y_clean)

        return pearson_corr, pearson_p, len(x_clean)
    except Exception as e:
        return None


def get_significance_level(p_value):

    if p_value < 0.001:
        return "***"
    elif p_value < 0.01:
        return "**"
    elif p_value < 0.05:
        return "*"
    else:
        return "ns"


def get_interpretation(pearson_corr, sig_level):

    abs_corr = abs(pearson_corr)

    if abs_corr < 0.3:
        strength = "Weak"
    elif abs_corr < 0.7:
        strength = "Moderate"
    else:
        strength = "Strong"

    direction = "positive" if pearson_corr >= 0 else "negative"

    return f"{strength} {direction} correlation ({sig_level})"


def load_data(file_paths):

    dfs = {}

    for name, file_path in file_paths.items():
        if not os.path.exists(file_path):
            print(f" :  {file_path}")
            sys.exit(1)

        try:
            df = pd.read_csv(file_path)

            print(f"\n {name}:")
            print(f"  : {len(df)} , {len(df.columns)} ")


            all_model_names = set(df['model'].tolist())


            df['Difficulty'] = df['model'].apply(lambda x: extract_difficulty_perfect(x, all_model_names))


            df_normalized = normalize_metrics(df, name)


            if name == 'Sonarqube':
                print(f"   Sonarqube:")
                df_normalized = add_sonarqube_aggregated_metrics(df_normalized)

            dfs[name] = df_normalized
            print(f"   : {len(df_normalized)} , {len(df_normalized.columns)} ")

        except Exception as e:
            print(f" :  {file_path}")
            print(f"   {str(e)}")
            sys.exit(1)

    return dfs


def calculate_all_correlations(dfs, difficulties):

    all_results = []
    exclude_cols = ['model', 'Total Codes', 'Correctness Rate (%)', 'Difficulty']

    print("\n" + "=" * 80)
    print(" Correctness RatePearson")
    print("    ")
    print("=" * 80)

    for difficulty in difficulties:
        print(f"\n: {difficulty}")

        for tool_name, df in dfs.items():

            df_filtered = df[df['Difficulty'] == difficulty].copy()

            if len(df_filtered) < 3:
                print(f"  {tool_name}:  ({len(df_filtered)})")
                continue

            print(f"  {tool_name}: {len(df_filtered)} ", end="")


            y = df_filtered['Correctness Rate (%)'].values


            all_cols = df.columns.tolist()


            processed_metrics = set()
            metric_count = 0


            metrics_to_norm = METRICS_TO_NORMALIZE.get(tool_name, [])

            for col in all_cols:
                if col in exclude_cols:
                    continue


                base_metric = col.replace(' (per code)', '')


                if base_metric in processed_metrics:
                    continue


                if base_metric in metrics_to_norm and f"{base_metric} (per code)" in all_cols:

                    metric = f"{base_metric} (per code)"
                elif ' (per code)' in col:

                    metric = col
                else:

                    metric = col


                processed_metrics.add(base_metric)

                if metric not in df_filtered.columns:
                    continue

                x = df_filtered[metric].values


                result = calculate_correlation(x, y)

                if result is None:
                    continue

                pearson_corr, pearson_p, sample_size = result


                sig_level = get_significance_level(pearson_p)


                direction = "Positive" if pearson_corr >= 0 else "Negative"


                interpretation = get_interpretation(pearson_corr, sig_level)


                pearson_score = abs(pearson_corr) * 100

                all_results.append({
                    'Difficulty': difficulty,
                    'Tool': tool_name,
                    'Metric': metric,
                    'Pearson Score': pearson_score,
                    'Direction': direction,
                    'Significance': sig_level,
                    'Pearson Corr': pearson_corr,
                    'P-value': pearson_p,
                    'Interpretation': interpretation,
                    'Sample Size': sample_size,
                    'Is Normalized': '(per code)' in metric
                })

                metric_count += 1

            print(f"   {metric_count} ")

    print(f"\n  {len(all_results)} ")

    return pd.DataFrame(all_results)


def create_excel_workbook(results_df, difficulties, output_path):

    print("\n" + "=" * 80)
    print("  + ")
    print("=" * 80)


    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for difficulty in difficulties:
        df_diff = results_df[results_df['Difficulty'] == difficulty].copy()
        df_diff = df_diff.sort_values('Pearson Score', ascending=False).reset_index(drop=True)

        print(f"\n{difficulty}: {len(df_diff)} ")


        normalized_count = df_diff['Is Normalized'].sum()
        print(f"  - : {normalized_count}/{len(df_diff)}")


        ws = wb.create_sheet(f'Ranking_{difficulty}')


        color_map = {
            'Easy': "70AD47",
            'Medium': "FFC000",
            'Hard': "C00000",
            'All': "4472C4"
        }
        header_color = color_map.get(difficulty, "4472C4")


        title_cell = ws.cell(1, 1)
        title_cell.value = f"Correctness Rate Correlation Ranking - Difficulty: {difficulty} (Enhanced with Aggregated Metrics)"
        title_cell.font = Font(bold=True, size=14, color=header_color)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)


        headers = [
            'Rank', 'Tool', 'Metric', 'Normalized', 'Pearson Score', 'Direction',
            'Significance', 'Pearson Corr', 'P-value',
            'Sample Size', 'Interpretation'
        ]

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(2, col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border


        for i, row in df_diff.iterrows():
            rank = i + 1
            excel_row = i + 3


            cell = ws.cell(excel_row, 1, rank)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

            if rank <= 5:
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                cell.font = Font(bold=True)
            elif rank <= 10:
                cell.fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")


            cell = ws.cell(excel_row, 2, row['Tool'])
            cell.border = thin_border


            cell = ws.cell(excel_row, 3, row['Metric'])
            cell.border = thin_border


            cell = ws.cell(excel_row, 4, "" if row['Is Normalized'] else "")
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if row['Is Normalized']:
                cell.font = Font(color="00B050")


            cell = ws.cell(excel_row, 5, round(row['Pearson Score'], 2))
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            cell.font = Font(bold=True)


            cell = ws.cell(excel_row, 6, row['Direction'])
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if row['Direction'] == "Positive":
                cell.font = Font(color="00B050")
            else:
                cell.font = Font(color="C00000")


            cell = ws.cell(excel_row, 7, row['Significance'])
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            cell.font = Font(bold=True)


            cell = ws.cell(excel_row, 8, round(row['Pearson Corr'], 4))
            cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border


            cell = ws.cell(excel_row, 9, row['P-value'])
            cell.number_format = '0.00E+00'
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border


            cell = ws.cell(excel_row, 10, int(row['Sample Size']))
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border


            cell = ws.cell(excel_row, 11, row['Interpretation'])
            cell.border = thin_border


        column_widths = {
            1: 6,
            2: 12,
            3: 45,
            4: 10,
            5: 13,
            6: 10,
            7: 12,
            8: 13,
            9: 12,
            10: 12,
            11: 40
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width


        ws.freeze_panes = 'A3'


    ws_info = wb.create_sheet('Info', 0)

    ws_info.cell(1, 1, " + Sonarqube").font = Font(bold=True, size=14)

    info_text = [
        "",
        "",
        "",
        " Pylint ",
        " Total Issues  Total Issues (per code)",
        " Convention (C)  Convention (C) (per code)",
        " Refactor (R)  Refactor (R) (per code)",
        " Warning (W)  Warning (W) (per code)",
        " Error (E)  Error (E) (per code)",
        " Fatal (F)  Fatal (F) (per code)",
        "",
        " Bandit ",
        " Codes With Issues  Codes With Issues (per code)",
        " Total Issues  Total Issues (per code)",
        " High/Medium/Low Severity   Severity (per code)",
        " High/Medium/Low Confidence   Confidence (per code)",
        "",
        " Radon ",
        " Total Complexity  Total Complexity (per code)",
        " Total LOC/LLOC/SLOC   LOC (per code)",
        " Total Comments/Blank  Comments/Blank (per code)",
        " Total Functions  Total Functions (per code)",
        " Grade A/B/C/D/E/F   Grade (per code)",
        "",
        " Sonarqube ",
        " Total Issues  Total Issues (per code)",
        " Severity: Blocker/Critical/Major/Minor/Info   Severity (per code)",
        " Type: Bug/Vulnerability/Code Smell/Security Hotspot   Type (per code)",
        " SQ Security/Reliability/Maintainability    (per code)",
        "",
        " Sonarqube ",
        " Avg Reliability Issues Per Code",
        "  - SQ Reliability: Blocker/High/Medium/Low/Info",
        " Avg Maintainability Issues Per Code",
        "  - SQ Maintainability: Blocker/High/Medium/Low/Info",
        " Avg Security Issues Per Code",
        "  - SQ Security: Blocker/High/Medium/Low/Info",
        "",
        "/",
        "  Avg ",
        "  (%) ",
        " Issue Rate ",
        "",
        "model",
        "'(per code)'",
        ""
    ]

    for i, text in enumerate(info_text, 2):
        ws_info.cell(i, 1, text)
        if text.startswith(""):
            ws_info.cell(i, 1).font = Font(color="0066CC")
        elif text.startswith("") or text.startswith("") or text.startswith("") or text.startswith(
                "") or text.startswith(""):
            ws_info.cell(i, 1).font = Font(bold=True, color="000080")

    ws_info.column_dimensions['A'].width = 80


    wb.save(output_path)

    print("\n" + "=" * 80)
    print(" Correctness Rate")
    print(f" : {output_path}")
    print("\nsheets:")
    print("   Info           -  + ")
    print("   Ranking_Easy   - ")
    print("   Ranking_Medium - ")
    print("   Ranking_Hard   - ")
    print("   Ranking_All    - ")
    print("=" * 80)


def print_summary(results_df, difficulties):

    print("\n" + "=" * 100)
    print(" Correctness Rate  - Top 5 ")
    print("=" * 100)

    emoji_map = {
        'Easy': "",
        'Medium': "",
        'Hard': "",
        'All': ""
    }

    for difficulty in difficulties:
        df_diff = results_df[results_df['Difficulty'] == difficulty].copy()
        df_diff = df_diff.sort_values('Pearson Score', ascending=False).head(5)

        emoji = emoji_map.get(difficulty, "")

        print(f"\n{emoji} {difficulty.upper()} - Top 5:")
        print("-" * 100)

        for i, row in df_diff.iterrows():
            rank = list(df_diff.index).index(i) + 1
            direction_emoji = "" if row['Direction'] == "Positive" else ""
            normalized_mark = "" if row['Is Normalized'] else "  "
            print(f"  {rank}. {normalized_mark} {row['Tool']:<12} - {row['Metric']:<45} "
                  f"(: {row['Pearson Score']:>6.2f}, {direction_emoji} {row['Direction']}, {row['Significance']})")

    print("\n =  (per code) ")
    print("\n" + "=" * 100)


def print_difficulty_distribution(dfs):

    print("\n" + "=" * 80)
    print(" ")
    print("=" * 80)

    for name, df in dfs.items():
        print(f"\n{name}:")
        dist = df['Difficulty'].value_counts().to_dict()

        for difficulty in ['All', 'Easy', 'Medium', 'Hard']:
            count = dist.get(difficulty, 0)
            expected = 73
            status = "" if count == expected else ""
            print(f"  {difficulty:6s}: {count:3d} model {status} (: {expected})")


def print_normalization_summary(dfs):

    print("\n" + "=" * 80)
    print(" ")
    print("=" * 80)

    for tool_name, df in dfs.items():
        print(f"\n{tool_name}:")


        normalized_cols = [col for col in df.columns if '(per code)' in col]
        original_metrics = METRICS_TO_NORMALIZE.get(tool_name, [])
        already_normalized = METRICS_ALREADY_NORMALIZED.get(tool_name, [])

        print(f"   : {len(df.columns) - len(normalized_cols) - 1}")
        print(f"   : {len(normalized_cols)}")
        print(f"   : {len(df.columns)}")
        print(f"   : {len(original_metrics)}")
        print(f"   /: {len(already_normalized)}")


        if tool_name == 'Sonarqube':
            aggregated_metrics = [
                'Avg Reliability Issues Per Code',
                'Avg Maintainability Issues Per Code',
                'Avg Security Issues Per Code'
            ]
            present_aggregated = [m for m in aggregated_metrics if m in df.columns]
            print(f"   : {len(present_aggregated)}")


def main():

    print("=" * 80)
    print("Correctness Rate  ( v6.1 - )")
    print("=" * 80)
    print(":")
    print("   ")
    print("   Pylint: 6")
    print("   Bandit: 9")
    print("   Radon: 13")
    print("   Sonarqube: 29")
    print("   model")
    print("   Sonarqube3")
    print("     - Avg Reliability Issues Per Code")
    print("     - Avg Maintainability Issues Per Code")
    print("     - Avg Security Issues Per Code")
    print("=" * 80)
    print()


    script_dir = os.path.dirname(os.path.abspath(__file__))

    file_paths = {
        'Bandit': os.path.join(script_dir, 'bandit_summary.csv'),
        'Pylint': os.path.join(script_dir, 'pylint_summary.csv'),
        'Radon': os.path.join(script_dir, 'radon_summary.csv'),
        'Sonarqube': os.path.join(script_dir, 'sonarqube_summary.csv')
    }


    output_path = os.path.join(script_dir, 'Correctness_Rate_Ranking_By_Difficulty_V6_Enhanced.xlsx')


    difficulties = ['Easy', 'Medium', 'Hard', 'All']


    print(" 1: ")
    print("-" * 80)
    dfs = load_data(file_paths)


    print_difficulty_distribution(dfs)


    print_normalization_summary(dfs)


    print("\n 2: Pearson")
    print("-" * 80)
    results_df = calculate_all_correlations(dfs, difficulties)


    print("\n 3: Excel")
    print("-" * 80)
    create_excel_workbook(results_df, difficulties, output_path)


    print_summary(results_df, difficulties)

    print("\n ")
    print(f"\n: {output_path}")


if __name__ == "__main__":
    main()